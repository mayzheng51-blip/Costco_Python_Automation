from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]   # Costco_Python_Automation/
MODEL = ROOT / "models" / "Costco – Elite Financial Model & Valuation.xlsx"

wb = load_workbook(MODEL)
print("Workbook loaded successfully!")
print("Sheets:")
for sheet in wb.sheetnames:
    print(sheet)
    