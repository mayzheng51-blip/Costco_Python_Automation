from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MODEL = ROOT / "Models" / "Costco – Elite Financial Model & Valuation.xlsx"
OUT = ROOT / "outputs"

SCENARIO_SHEET = "Executive Dashboard"
SCENARIO_CELL = "E4"   # your scenario selector

def set_scenario_and_save(scenario: str):
    wb = load_workbook(MODEL)
    ws = wb[SCENARIO_SHEET]

    ws[SCENARIO_CELL].value = scenario

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUT / f"model_{scenario}_{ts}.xlsx"
    wb.save(out_path)

    print(f"Saved: {out_path.name}")

if __name__ == "__main__":
    OUT.mkdir(exist_ok=True)
    for s in ["Bear", "Base", "Bull"]:
        set_scenario_and_save(s)